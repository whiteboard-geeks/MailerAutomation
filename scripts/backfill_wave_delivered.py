#!/usr/bin/env python
"""
Backfill 'Mailer Delivered' updates in Close from a Pirate Ship delivery report.

Use this when EasyPost has stopped polling a batch of trackers but the packages
have actually been delivered (per Pirate Ship's spreadsheet). Performs the same
two writes the EasyPost delivery webhook would have done:
  1. Updates delivery custom fields on the lead
  2. Creates a 'Mailer Delivered' custom activity on the lead

Defaults to dry-run. Pass --apply to write to Close.

Usage:
  python -m scripts.backfill_wave_delivered <xlsx>            # dry-run
  python -m scripts.backfill_wave_delivered <xlsx> --apply    # write
  python -m scripts.backfill_wave_delivered <xlsx> --limit 3  # first 3 only
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env")

from openpyxl import load_workbook  # noqa: E402

from close_utils import (  # noqa: E402
    load_query,
    search_close_leads,
    update_delivery_information_for_lead,
)
from utils.easypost import create_package_delivered_custom_activity_in_close  # noqa: E402


# 1-based Pirate Ship column positions
COL_TRACKING = 4
COL_TRACK_DT = 18
COL_CITY = 21
COL_STATE = 22


def parse_tracking_date(s: str) -> datetime:
    """Parse "5/2/26 4:40 PM EDT" → datetime (drops trailing tz abbreviation)."""
    s = s.strip()
    head, _, tail = s.rpartition(" ")
    if tail.isalpha():
        s = head
    return datetime.strptime(s, "%m/%d/%y %I:%M %p")


def build_delivery_information(city: str, state: str, dt: datetime) -> dict:
    """Match temporal/activities/easypost/webhook_delivery_status._parse_delivery_information."""
    info: dict = {
        "delivery_city": city.title() if city else "N/A",
        "delivery_state": state.upper() if state else "N/A",
        "delivery_date": dt.date(),
        "delivery_date_readable": dt.strftime("%a %-m/%-d"),
    }
    info["date_and_location_of_mailer_delivered"] = (
        f"{info['delivery_date_readable']} to {info['delivery_city']}, {info['delivery_state']}"
    )
    info["location_delivered"] = f"{info['delivery_city']}, {info['delivery_state']}"
    return info


def find_lead_ids_by_tracking(tracking_code: str) -> list[str]:
    q = load_query("lead_by_tracking_number.json")
    q["query"]["queries"][1]["queries"][0]["queries"][0]["condition"]["value"] = tracking_code
    leads = search_close_leads(q)
    return [lead["id"] for lead in leads]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="Path to Pirate Ship delivered mailers xlsx")
    ap.add_argument("--apply", action="store_true", help="Write to Close (default: dry-run)")
    ap.add_argument("--limit", type=int, default=0, help="Stop after N rows (0 = no limit)")
    ap.add_argument("--start-row", type=int, default=2, help="First data row, 1-based (default 2)")
    args = ap.parse_args()

    if not os.environ.get("CLOSE_API_KEY"):
        sys.exit("CLOSE_API_KEY missing — make sure .env is loaded")

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb.active

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"=== {mode} | source={args.xlsx} | rows from {args.start_row} ===")

    counts = {
        "would_apply": 0,
        "fields_updated": 0,
        "activity_created": 0,
        "activity_skipped_exists": 0,
        "not_found": 0,
        "ambiguous": 0,
        "skipped_bad_row": 0,
        "error": 0,
    }

    processed = 0
    for i, row in enumerate(
        ws.iter_rows(min_row=args.start_row, values_only=True), start=args.start_row
    ):
        if args.limit and processed >= args.limit:
            break
        if not any(row):
            continue
        processed += 1

        recipient = row[0]
        tracking = row[COL_TRACKING - 1]
        tracking = str(tracking).strip() if tracking is not None else ""
        city = row[COL_CITY - 1] or ""
        state = row[COL_STATE - 1] or ""
        tdate = row[COL_TRACK_DT - 1]

        if not tracking or not tdate:
            print(f"row {i}: SKIP missing tracking or date  ({recipient})")
            counts["skipped_bad_row"] += 1
            continue

        try:
            dt = tdate if isinstance(tdate, datetime) else parse_tracking_date(str(tdate))
        except Exception as e:
            print(f"row {i}: ERROR parse date {tdate!r}  {e}")
            counts["error"] += 1
            continue

        try:
            lead_ids = find_lead_ids_by_tracking(tracking)
        except Exception as e:
            print(f"row {i}: ERROR search {tracking}  {e}")
            counts["error"] += 1
            continue

        if not lead_ids:
            print(f"row {i}: NOT_FOUND  {tracking}  ({recipient})")
            counts["not_found"] += 1
            continue
        if len(lead_ids) > 1:
            print(f"row {i}: AMBIGUOUS  {tracking}  {len(lead_ids)} matches  ({recipient})")
            counts["ambiguous"] += 1
            continue

        lead_id = lead_ids[0]
        info = build_delivery_information(city, state, dt)
        summary = info["date_and_location_of_mailer_delivered"]

        if not args.apply:
            print(f"row {i}: WOULD_APPLY  {tracking}  lead={lead_id}  {summary}  ({recipient})")
            counts["would_apply"] += 1
            continue

        try:
            update_delivery_information_for_lead(lead_id, info)
            counts["fields_updated"] += 1
        except Exception as e:
            print(f"row {i}: ERROR update lead {lead_id}  {e}")
            counts["error"] += 1
            continue

        try:
            resp = create_package_delivered_custom_activity_in_close(lead_id, info)
            if resp.get("status") == "skipped":
                counts["activity_skipped_exists"] += 1
                marker = "ACT_EXISTS"
            else:
                counts["activity_created"] += 1
                marker = "ACT_CREATED"
        except Exception as e:
            print(f"row {i}: PARTIAL fields_updated but activity FAILED for {lead_id}  {e}")
            counts["error"] += 1
            continue

        print(f"row {i}: OK  {tracking}  lead={lead_id}  {marker}  {summary}  ({recipient})")

    print()
    print(f"=== summary ({mode}) ===")
    for k, v in counts.items():
        print(f"  {k:24s} {v}")


if __name__ == "__main__":
    main()
